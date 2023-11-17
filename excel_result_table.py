import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment, Color 
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from wcwidth import wcswidth

import os
import datetime
import re

import win32com.client

import pandas as pd
import sys

# 摘要用
# from summary import summary_result_table

def result_table(out_table: pd.DataFrame):
    
    if type(out_table)==type(pd.DataFrame()):

        source_df = out_table[out_table.DailyNews == "1"].drop_duplicates(['title'])

        # 日期轉換後，與資料來源合併
        source_df["datetimeParsed"] = pd.to_datetime(source_df.datetime, format='%Y-%m-%d %H:%M:%S').dt.strftime('%m/%d %H:%M')

        # 日期轉換後，與資料來源合併
        source_df["site_datetimeParsed"] = source_df.site + "\n" + pd.to_datetime(source_df.datetime, format='%Y-%m-%d %H:%M:%S').dt.strftime('%m/%d %H:%M')
        # 標題與連結合併
        source_df["title_link"] = source_df.title + "\n" + source_df.short_link

        # 取出所需欄位
        select_df = source_df.apply(lambda row: row[["category", "title_link", "site_datetimeParsed"]], axis=1)

        # 改欄位名稱
        columns = select_df.columns.to_list()
        new_columns = ["類別" , "新聞", "出處/日期"]
        select_df = select_df.rename(columns=dict(zip(columns, new_columns)))

        # 篩選出 '產業' 的前三筆
        industry_df = select_df[select_df['類別'] == '產業'].head(4)

        # 篩選出 '法令' 的前三筆
        law_df = select_df[select_df['類別'] == '法令'].head(2) 

        # 篩選出 'HR' 的前三筆
        hr_df = select_df[select_df['類別'] == 'HR'].head(2)  

        # 合併 DataFrame
        result_df = pd.concat([industry_df, law_df, hr_df])

        # 類別正名
        result_df.類別 = result_df.類別.apply(lambda s: "產業趨勢" if s=="產業" else "HR相關新聞" if s=="HR" else s)

        # 摘要
        title_list = result_df.新聞.str.extract("(.+)\n(http.+)")[0]
        link_list = result_df.新聞.str.extract("(.+)\n(http.+)")[1]
        # summary_df = summary_result_table(title_list, link_list)
        summary_df = pd.DataFrame()

        today = datetime.datetime.now().strftime("%Y%m%d")
        current_path = os.getcwd()
        filename = f"News_{today}.xlsx"
        path = f".\\history_result\\{filename}"
        abs_path = current_path + path[1:]


        # 如果檔案有開，先關檔案
        def close_specific_excel_file(file_path):
            if sys.platform == "win32":
                try:
                    xl = win32com.client.Dispatch('Excel.Application')
                    for wb in xl.Workbooks:
                        if wb.FullName.lower() == file_path.lower():
                            wb.Close(SaveChanges=True)
                            break
                    if len(xl.Workbooks) == 0:
                        xl.Quit()
                except Exception as e:
                    print(f"Error: {e}")

        close_specific_excel_file(abs_path)

        with pd.ExcelWriter(path, engine='openpyxl') as writer:

            wb = openpyxl.Workbook()
            wb.create_sheet(title="彙整表")
            wb.create_sheet(title="摘要")
            wb.remove(wb['Sheet']) 
            result_sheet = wb["彙整表"]
            summary_sheet = wb['摘要']
            result_sheet.sheet_properties.line_terminator = '\n' # 设置工作表的 line_terminator 属性为换行符（\n），用于分隔表格行

            # 把df的值寫入excel
            writer.book = wb
            result_df.to_excel(writer, sheet_name=result_sheet.title, index=False, startrow=0)
            summary_df.to_excel(writer, sheet_name=summary_sheet.title, index=False, header=False,startrow=0)

            ### 設定"彙整表"sheet樣式
            # 設定第一列樣式
            row_fixed = 1
            for c in range(result_sheet.min_column, result_sheet.max_column + 1):
                cell = result_sheet.cell(row=row_fixed, column=c)
                cell.font = Font(color='FFFFFF', bold=True)
                cell.fill = PatternFill(fgColor='006699', fill_type='solid')

            # 設定樣式與合併   
            def merge_cross_row_and_align(sheet, start_row, start_column, end_row, end_column, align=True, color='000000', bold=True):
                sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
                cell = sheet.cell(row=start_row, column=start_column)
                if align:
                    cell.alignment = Alignment(horizontal='center', vertical='center') 
                cell.font = Font(color='000000', bold=True)

            # 第一欄的樣式與合併  
            col_fixed = 1
            pre_category, start_row, stop_row = None, None, None
            for r in range(result_sheet.min_row+1, result_sheet.max_row + 1):
                category = result_sheet.cell(row=r, column=col_fixed).value

                if category != pre_category:
                    if pre_category is not None: 
                         merge_cross_row_and_align(result_sheet, start_row=start_row, start_column=col_fixed, end_row=stop_row, end_column=col_fixed)
                    start_row = r            

                pre_category = category
                stop_row = r

            if result_sheet.min_row+1> 1:
                 merge_cross_row_and_align(result_sheet, start_row=start_row, start_column=col_fixed, end_row=stop_row, end_column=col_fixed)


            # 設定第二欄的樣式與合併，提出短網址到short_link_list
            col_fixed = 2
            pattern = "(.+)\n(http.+)"
            for r in range(result_sheet.min_row+1, result_sheet.max_row + 1):
                cell = result_sheet.cell(row=r, column=col_fixed)
                string = cell.value
                title_str = re.search(pattern, string).group(1)
                link_str = re.search(pattern, string).group(2)

                new_string = f'=HYPERLINK("{link_str}", "{string}")'

                cell.value = new_string
                cell.font = Font(color='005087')


            # 设置字體、邊框   
            def set_auo_font(cell, size=12):
                original_font = cell.font
                text = str(cell.value)
                is_chinese = any("\u4E00" <= c <= "\u9FFF" for c in text) # 使用Unicode范围检查是否包含中文字符
                try:
                    if is_chinese:
                        cell.font = Font(name="Noto Sans TC", size=size, color=original_font.color, bold=original_font.bold, italic=original_font.italic)
                    else:
                        cell.font = Font(name="Mabry Pro", size=size, color=original_font.color, bold=original_font.bold, italic=original_font.italic)
                except Exception as e:
                    print(e)

            border = Border(
                left   = Side(border_style="thin"),
                right  = Side(border_style="thin"), 
                top    = Side(border_style="thin"),
                bottom = Side(border_style="thin")
            )

            scale_factor = 1.3 # 调整合适的缩放因子
            result_df_width = 0
            for c in range(result_sheet.min_column, result_sheet.max_column + 1):
                max_width = 0
                column_letter = get_column_letter(c)
                for r in range(result_sheet.min_row, result_sheet.max_row + 1):
                    cell = result_sheet.cell(row=r, column=c)
                    # 設字體
                    set_auo_font(cell, size=12)
                    # # 为cell加边框
                    cell.border = border
                    if '\n' in str(cell.value):
                        cell.alignment = Alignment(wrap_text=True)

                    need_width = max(wcswidth(line) for line in str(cell.value).split("\n"))

                    if need_width > max_width:
                        max_width = need_width

                # 根据宽度设置的缩放因子
                result_sheet.column_dimensions[column_letter].width = max_width * scale_factor
                result_df_width += result_sheet.column_dimensions[column_letter].width 

            ### 設定"摘要"sheet樣式
            col_fixed = 1
            for r in range(summary_sheet.min_row, summary_sheet.max_row + 1):
                cell = summary_sheet.cell(row=r, column=col_fixed)
                if r%2:
                    cell.font = Font(color='005087', bold=True)
                    set_auo_font(cell, size=14)
                else:
                    set_auo_font(cell, size=12)

            # 調跟result_df一樣寬
            summary_sheet.column_dimensions[get_column_letter(1)].width = result_df_width

        # 開啟檔案
        os.system(f'start excel.exe {abs_path}') 
        
        print("已將表格輸出至history_tag、history_result資料夾")
        
        return result_df